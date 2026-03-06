#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import logging, glob, re, hashlib, datetime
from pathlib import Path
import pandas as pd
from openpyxl.styles import Alignment, Font
from tqdm.auto import tqdm

# ── CONFIG ───────────────────────────────────────────────────────────────
DATA_DIR = Path("/Users/nickbowditch/Documents/PHD/DATA")
IN_PARQ  = DATA_DIR / "stage3_predicted.parquet"
MANIFEST = DATA_DIR / "manifest.csv"

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
pd.options.mode.copy_on_write = True

# ── Helpers ──────────────────────────────────────────────────────────────
def next_version() -> float:
    files = glob.glob(str(DATA_DIR / "participant_data_schemaV8_v*_golden.xlsx"))
    if not files:
        return 1.0
    vs = []
    for f in files:
        m = re.search(r"_v(\d+\.\d+)_golden", f)
        if m:
            vs.append(float(m.group(1)))
    return round(max(vs) + 0.1, 1) if vs else 1.0

# ── MAIN ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if not IN_PARQ.exists():
        raise FileNotFoundError(f"Missing {IN_PARQ}")

    df = pd.read_parquet(IN_PARQ)
    logging.info(f"✅ Loaded {IN_PARQ.name} ({df.shape[0]} rows × {df.shape[1]} cols)")

    ver = next_version()
    out_xlsx = DATA_DIR / f"participant_data_schemaV8_v{ver:.1f}_golden.xlsx"

    # Excel export with progress bar
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    # Write header
    ws.append(list(df.columns))

    # Write rows with progress bar
    with tqdm(total=len(df), desc="Excel Export", unit="row") as pbar:
        for _, row in df.iterrows():
            ws.append(list(row.values))
            pbar.update(1)

    # Format cells (wrap + font) with progress bar on visible rows
    from openpyxl.utils import get_column_letter
    with tqdm(total=ws.max_row, desc="Formatting", unit="row") as pbar2:
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in r:
                cell.alignment = Alignment(wrap_text=True)
                cell.font = Font(name="Helvetica Neue", size=12)
            pbar2.update(1)

    # Set column widths: L & AF = 50, others = 15
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 100 if letter in ("L", "AF") else 20

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(out_xlsx)
    logging.info(f"✅ Saved {out_xlsx.name}")

    # ── Manifest update (consistent with earlier format) ──────────────────
    h = hashlib.md5(pd.util.hash_pandas_object(df, index=True).values).hexdigest()
    with open(MANIFEST, "a") as mf:
        mf.write(f"v8_final,{IN_PARQ.name},{out_xlsx.name},{len(df)},{h},{datetime.datetime.now()}\n")
    logging.info(f"📜 Manifest updated → {MANIFEST.name}")

    logging.info("🏁 Stage 4 complete.")