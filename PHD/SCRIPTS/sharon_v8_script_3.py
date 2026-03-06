\#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import pandas as pd, numpy as np, joblib, logging, re, datetime, hashlib, glob, os
from pathlib import Path
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────
DATA_DIR   = Path("/Users/nickbowditch/Documents/PHD/DATA")
FLAT_IN    = DATA_DIR / "sharon_v8_flattened.parquet"
MODEL_PATH = Path("/Users/nickbowditch/Documents/PHD/MODELS/sharon_v14.6.pkl")
MANIFEST   = DATA_DIR.parent / "manifest.csv"
BACKUP_PAR = DATA_DIR / "sharon_v8_script3_full.parquet"

# Locked schema order
ORDER = """participant_id	age	gender	indigenous_status	income_level	education_level	postcode	rural_remote_status	dropout_actual	relapse_actual	dropout_prediction	relapse_prediction	narrative_dropout	narrative_relapse	explainability_report	dropout_prediction_probability	relapse_prediction_probability	dropout_prediction_calibrated	relapse_prediction_calibrated	dropout_prediction_confidence	relapse_prediction_confidence	dropout_trajectory	relapse_trajectory	trajectory_risk_dropout	trajectory_risk_relapse	z_score_dropout	z_score_relapse	adaptive_risk_dropout	adaptive_risk_relapse	transcript_type	transcript	DTCQ-8	URICA-S	BAM-R	sentiment_score	fused_sentiment_score	sentiment_change	fused_sentiment_change	sentiment_slope	context_sentiment_avg	word_count	fused_word_count	first_person_ratio	fused_first_person_ratio	lexical_diversity	fused_lexical_diversity	Happy	prev_Happy_avg	slope_Happy	Angry	prev_Angry_avg	slope_Angry	Surprise	prev_Surprise_avg	slope_Surprise	Sad	prev_Sad_avg	slope_Sad	Fear	prev_Fear_avg	slope_Fear	engagement_decay	fused_engagement_decay	reflective_language_score	fused_reflective_language_score	text_length	text_length_change	fused_text_length_change	slope_text_length_change	screen_unlocks_per_day	total_unlocks	estimated_sleep_duration	sleep_time	wake_time	sleep_duration_hours	late_night_minutes	late_night_screen_usage_YN	notifications_delivered	notifications_opened	notification_engagement_rate	notifi_latency_avg_s	notifi_latency_p50_s	notifi_latency_p90_s	notifi_latency_p99_s	notifi_latency_count	daily_usage_entropy_bits	usage_event_count	app_min_total	app_min_social	app_min_dating	app_min_productivity	app_min_music_audio	app_min_image	app_min_maps	app_min_video	app_min_travel_local	app_min_shopping	app_min_news	app_min_game	app_min_health	app_min_finance	app_min_browser	app_min_comm	app_min_other	app_switches	app_switch_entropy	distance_km	movement_intensity	screen_usage_min	notification_count	app_usage_min	""" + "\t".join([f"embed{str(i).zfill(3)}" for i in range(1,769)])
ORDER = [c for c in ORDER.split("\t") if c]

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

# ─── helpers ─────────────────────────────────────────────────────────────
def latest_v8_golden_version() -> float:
    cands = glob.glob(str(DATA_DIR / "participant_data_schemaV8_v*_golden.xlsx")) + \
            glob.glob(str(DATA_DIR / "participant_data_schemaV8_v*_golden.parquet"))
    if not cands: return 1.0
    def ver(p):
        m = re.search(r"_v(\d+\.\d+)_golden", os.path.basename(p))
        return float(m.group(1)) if m else 1.0
    return max(map(ver, cands))

def next_v8_version(cur: float) -> float:
    return round(cur + 0.1, 1)

META_COLS = {
    "participant_id","transcript_type","transcript",
    "dropout_actual","relapse_actual",
    "age","gender","indigenous_status","income_level","education_level",
    "postcode","rural_remote_status"
}

def flatten(obj):
    seen=set(); stack=[obj]
    while stack:
        x=stack.pop()
        if id(x) in seen: 
            continue
        seen.add(id(x)); yield x
        # common containers
        for attr in ("dropout","relapse","models","best_estimator_","base_estimator","calibrated_classifiers_"):
            if hasattr(x, attr):
                v=getattr(x, attr)
                if isinstance(v,(list,tuple)): stack.extend(v)
                elif isinstance(v,dict): stack.extend(v.values())
                elif v is not None: stack.append(v)
        for attr in ("steps","named_steps","transformers_","estimators_"):
            if hasattr(x, attr):
                v=getattr(x, attr)
                if isinstance(v, list):
                    for item in v:
                        if isinstance(item,(list,tuple)):
                            for p in item:
                                if not isinstance(p,str): stack.append(p)
                        else:
                            stack.append(item)
                elif isinstance(v, dict):
                    stack.extend(v.values())

def extract_feature_names(model, df):
    # 1) feature_names_in_ anywhere
    for node in flatten(model):
        fni=getattr(node,"feature_names_in_",None)
        if fni is not None and len(fni)>0:
            cols=[c for c in list(fni) if c in df.columns]
            if cols: return cols, "feature_names_in_"
    # 2) get_feature_names_out (ColumnTransformer, etc.)
    for node in flatten(model):
        g=getattr(node,"get_feature_names_out",None)
        if callable(g):
            try:
                cols=[c.split("__")[-1] for c in list(g())]
                cols=[c for c in cols if c in df.columns]
                if cols: return cols, "get_feature_names_out"
            except Exception:
                pass
    # 3) embeddings heuristic
    embed_cols=[c for c in df.columns if re.fullmatch(r"embed\d{3}", str(c))]
    if len(embed_cols)>=64: return embed_cols, "heuristic: embeddings"
    # 4) numeric fallback
    num_cols=[c for c in df.columns if c not in META_COLS and pd.api.types.is_numeric_dtype(df[c])]
    if len(num_cols)>=5: return num_cols, "fallback: numeric"
    return [], "none"

def proba(est, X):
    P = est.predict_proba(X)
    return P[:,1] if getattr(P,"ndim",2)==2 else P

def slope_nan_robust(vals: np.ndarray) -> float:
    y = np.asarray(vals, float)
    mask = ~np.isnan(y)
    y = y[mask]
    if y.size < 2: return 0.0
    x = np.arange(y.size, dtype=float)
    x -= x.mean(); y -= y.mean()
    d = (x**2).sum()
    return float((x*y).sum()/d) if d else 0.0

def slope_to_traj(m: float, eps: float=0.01) -> str:
    if m > eps: return "rising"
    if m < -eps: return "falling"
    return "stable"

def bucket_or_blank(p):
    if pd.isna(p): return ""
    if p >= 0.80: return "very_high"
    if p >= 0.60: return "high"
    if p >= 0.40: return "moderate"
    if p >= 0.20: return "low"
    return "very_low"

def safe_float(v, default=0.0):
    try:
        if pd.isna(v): return float(default)
        return float(v)
    except Exception:
        return float(default)

# ─── load ────────────────────────────────────────────────────────────────
df = pd.read_parquet(FLAT_IN)
model = joblib.load(MODEL_PATH)
logging.info(f"✅ Data {df.shape}, model loaded.")

# ─── features ────────────────────────────────────────────────────────────
features, feat_rule = extract_feature_names(model, df)
if not features:
    raise SystemExit("✖ Could not determine feature list for prediction.")
logging.info(f"✅ Feature discovery: {feat_rule}; using {len(features)} columns.")

# rows with complete features
missing_mask = df[features].isna().any(axis=1)
ok_mask      = ~missing_mask
logging.info(f"✅ Rows with complete features: {int(ok_mask.sum())}; rows needing features: {int(missing_mask.sum())}")

# model handles
md = model["dropout"] if isinstance(model, dict) else model.dropout
mr = model["relapse"] if isinstance(model, dict) else model.relapse

# make prediction columns with safe dtypes (prevents FutureWarning)
for c, dtype in [
    ("dropout_prediction_probability", "float64"),
    ("relapse_prediction_probability", "float64"),
    ("dropout_prediction_calibrated",  "float64"),
    ("relapse_prediction_calibrated",  "float64"),
    ("dropout_prediction", "string"),
    ("relapse_prediction", "string"),
    ("dropout_prediction_confidence", "float64"),
    ("relapse_prediction_confidence", "float64"),
]:
    if c not in df.columns:
        df[c] = pd.Series(index=df.index, dtype=dtype)

# predict only on complete rows
if int(ok_mask.sum()) > 0:
    X_ok = df.loc[ok_mask, features].astype(float).to_numpy()
    pdrop = proba(md, X_ok)
    prel  = proba(mr, X_ok)

    df.loc[ok_mask, "dropout_prediction_probability"] = pdrop
    df.loc[ok_mask, "relapse_prediction_probability"] = prel
    df.loc[ok_mask, "dropout_prediction_calibrated"]  = pdrop
    df.loc[ok_mask, "relapse_prediction_calibrated"]  = prel

    def label_from(p): return np.where(p>=0.5, "Y", "N")
    def conf_from(p):  return np.abs(p-0.5)*2.0

    # cast to string explicitly → no dtype warnings
    df.loc[ok_mask, "dropout_prediction"] = pd.Series(label_from(df.loc[ok_mask, "dropout_prediction_calibrated"].to_numpy()), index=df.index[ok_mask], dtype="string")
    df.loc[ok_mask, "relapse_prediction"] = pd.Series(label_from(df.loc[ok_mask, "relapse_prediction_calibrated"].to_numpy()), index=df.index[ok_mask], dtype="string")
    df.loc[ok_mask, "dropout_prediction_confidence"] = conf_from(df.loc[ok_mask, "dropout_prediction_calibrated"].to_numpy())
    df.loc[ok_mask, "relapse_prediction_confidence"] = conf_from(df.loc[ok_mask, "relapse_prediction_calibrated"].to_numpy())

# status column (for rows we skipped)
status_col = "prediction_status"
if status_col not in df.columns:
    df[status_col] = ""
df.loc[ok_mask,      status_col] = "ok"
df.loc[missing_mask, status_col] = "missing_embeddings"

# ─── trajectories & z-scores (NaN-robust) ───────────────────────────────
order_map = {"baseline":0,"voicenote1":1,"voicenote2":2,"voicenote3":3,"voicenote4":4,"voicenote5":5,
             "voicenote6":6,"voicenote7":7,"voicenote8":8,"voicenote9":9,"voicenote10":10,"voicenote11":11,"exit":12}
df["_tnum"] = df.get("transcript_type","").map(order_map).fillna(9999).astype(int)
df = df.sort_values(["participant_id","_tnum"], kind="stable")

traj_d, traj_r, z_d, z_r = [], [], [], []
for _, g in df.groupby("participant_id", sort=False):
    sd = pd.to_numeric(g["dropout_prediction_calibrated"], errors="coerce").to_numpy()
    sr = pd.to_numeric(g["relapse_prediction_calibrated"],  errors="coerce").to_numpy()

    md_ = slope_nan_robust(sd)
    mr_ = slope_nan_robust(sr)
    traj_d += [slope_to_traj(md_)]*len(g)
    traj_r += [slope_to_traj(mr_)]*len(g)

    # z-scores guarded (no warnings on all-NaN)
    if np.isfinite(sd).any():
        mu_d = float(np.nanmean(sd)); sig_d = float(np.nanstd(sd, ddof=0) + 1e-8)
        z_d  += ((sd - mu_d) / sig_d).tolist()
    else:
        z_d  += [0.0]*len(g)

    if np.isfinite(sr).any():
        mu_r = float(np.nanmean(sr)); sig_r = float(np.nanstd(sr, ddof=0) + 1e-8)
        z_r  += ((sr - mu_r) / sig_r).tolist()
    else:
        z_r  += [0.0]*len(g)

df["dropout_trajectory"] = traj_d
df["relapse_trajectory"] = traj_r
df["z_score_dropout"]    = z_d
df["z_score_relapse"]    = z_r
df["trajectory_risk_dropout"] = df["dropout_prediction_calibrated"].map(bucket_or_blank)
df["trajectory_risk_relapse"] = df["relapse_prediction_calibrated"].map(bucket_or_blank)

# ─── narratives ──────────────────────────────────────────────────────────
def categorize(p):
    if p > 0.7: return "high"
    if p < 0.3: return "low"
    return "moderate"

def session_report(r):
    d = safe_float(r.get("dropout_prediction_calibrated", np.nan), 0.0)
    R = safe_float(r.get("relapse_prediction_calibrated",  np.nan), 0.0)
    text = (
        f"This week, SHARON estimated the participant’s dropout risk at {d:.2f} "
        f"({categorize(d)}) and relapse risk at {R:.2f} ({categorize(R)}). "
    )
    signals = []
    suggestions = []
    if safe_float(r.get("engagement_decay", 0)) > 0.5:
        signals.append("behavioural disengagement"); suggestions.append("monitor engagement patterns")
    if safe_float(r.get("sentiment_score", r.get("sentiment", 0))) < -0.3:
        signals.append("negative mood"); suggestions.append("track mood and affect")
    if safe_float(r.get("text_length_change", 0)) < -0.2:
        signals.append("shortened responses"); suggestions.append("check in on communication consistency")
    if safe_float(r.get("Sad", 0)) > 2:
        signals.append("heightened sadness"); suggestions.append("consider mood support interventions")

    text += ("Key warning signals included " + ", ".join(signals) + ". ") if signals else "No clear warning signals were detected this week. "
    sug = "; ".join(dict.fromkeys(suggestions)) if suggestions else "maintain regular check-ins"
    text += f"Clinician suggestions: {sug}."
    return text

if "explainability_report" not in df.columns:
    df["explainability_report"] = pd.Series(index=df.index, dtype="string")
df["explainability_report"] = df.apply(session_report, axis=1)

def exec_summary(full: pd.DataFrame) -> str:
    dvals = pd.to_numeric(full["dropout_prediction_calibrated"], errors="coerce").to_numpy()
    rvals = pd.to_numeric(full["relapse_prediction_calibrated"],  errors="coerce").to_numpy()
    def first_non_nan(a, default=0.0):
        idx = np.where(~np.isnan(a))[0]
        return float(a[idx[0]]) if idx.size>0 else default
    def last_non_nan(a, default=0.0):
        idx = np.where(~np.isnan(a))[0]
        return float(a[idx[-1]]) if idx.size>0 else default

    first_d, last_d = first_non_nan(dvals), last_non_nan(dvals)
    first_r, last_r = first_non_nan(rvals), last_non_nan(rvals)
    trend_d = "increasing" if last_d > first_d else "decreasing" if last_d < first_d else "stable"
    trend_r = "increasing" if last_r > first_r else "decreasing" if last_r < first_r else "stable"

    weeks = list(range(1, len(full) + 1))
    spikes_d = [f"week {w}" for w,v in zip(weeks, dvals) if not np.isnan(v) and v > 0.7]
    spikes_r = [f"week {w}" for w,v in zip(weeks, rvals) if not np.isnan(v) and v > 0.7]

    all_signals = []
    for _, row in full.iterrows():
        if safe_float(row.get("engagement_decay",0)) > 0.5:      all_signals.append("behavioural disengagement")
        if safe_float(row.get("sentiment_score", row.get("sentiment",0))) < -0.3: all_signals.append("negative mood")
        if safe_float(row.get("text_length_change",0)) < -0.2:   all_signals.append("shortened responses")
        if safe_float(row.get("Sad",0)) > 2:                     all_signals.append("heightened sadness")
    top = list(dict.fromkeys(all_signals))[:4] or ["none"]

    parts = [
        f"As you wrap up this participant’s journey over {len(full)} entries, "
        f"dropout risk moved from {first_d:.2f} to {last_d:.2f} ({trend_d} trend) and "
        f"relapse risk from {first_r:.2f} to {last_r:.2f} ({trend_r} trend)."
    ]
    if spikes_d or spikes_r:
        sd = ", ".join(spikes_d) if spikes_d else "none"
        sr = ", ".join(spikes_r) if spikes_r else "none"
        parts.append(f"Significant spikes occurred for dropout in {sd}; relapse in {sr}.")
    parts.append(f"The most recurring warning signals were {', '.join(top)}.")
    parts.append("Please see the individual session narratives above for details, and consider tailored interventions based on these patterns.")
    return " ".join(parts)

exit_mask = df["transcript_type"].astype(str).str.lower() == "exit"
groups = {pid: grp for pid, grp in df.groupby("participant_id")}
for idx in df[exit_mask].index:
    pid = df.at[idx, "participant_id"]
    df.at[idx, "explainability_report"] = exec_summary(groups[pid])

# ─── strict order ────────────────────────────────────────────────────────
present = set(df.columns)
ordered = [c for c in ORDER if c in present]
extras  = [c for c in df.columns if c not in ordered and c != "_tnum"]
df = df[ordered + extras]

# ─── backup parquet ──────────────────────────────────────────────────────
df.to_parquet(BACKUP_PAR, index=False)
logging.info(f"✅ Backed up to {BACKUP_PAR.name}")

# ─── bump golden (xlsx) — widths only (fast), no per-cell wrap to avoid hangs ─
cur = latest_v8_golden_version()
new_ver = next_v8_version(cur)
gold_xlsx = DATA_DIR / f"participant_data_schemaV8_v{new_ver:.1f}_golden.xlsx"
if gold_xlsx.exists():
    gold_xlsx = gold_xlsx.with_name(gold_xlsx.stem + "(1)" + gold_xlsx.suffix)

with pd.ExcelWriter(gold_xlsx, engine="openpyxl") as w:
    df.to_excel(w, index=False, sheet_name="data")
    ws = w.sheets["data"]
    ws.freeze_panes = "A2"
    # set width 15 for all, L=50
    from openpyxl.utils import get_column_letter
    for j in range(1, ws.max_column+1):
        col_letter = get_column_letter(j)
        ws.column_dimensions[col_letter].width = 50 if col_letter == "L" else 15
logging.info(f"✅ Golden bumped → {gold_xlsx.name}")

# ─── clinician subset (wrap all; width=15 except L=50) ───────────────────
CLIN = [
    "participant_id","transcript_type",
    "trajectory_risk_dropout","trajectory_risk_relapse",
    "dropout_prediction_calibrated","relapse_prediction_calibrated",
    "narrative_dropout","narrative_relapse",
    "explainability_report",
    "dropout_trajectory","relapse_trajectory",
    "dropout_prediction_probability","relapse_prediction_probability",
    "dropout_prediction_confidence","relapse_prediction_confidence",
    "z_score_dropout","adaptive_risk_dropout",
    "z_score_relapse","adaptive_risk_relapse",
    "dropout_actual","relapse_actual"
]
df_clin = df[[c for c in CLIN if c in df.columns]]
clin_xlsx = DATA_DIR / f"sharon_clinician_report_v{new_ver:.1f}.xlsx"

with pd.ExcelWriter(clin_xlsx, engine="openpyxl") as w:
    df_clin.to_excel(w, index=False, sheet_name="report")
    ws = w.sheets["report"]
    # wrap text on ALL cells + font (sheet is small → fast)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            cell.font = Font(name="Helvetica Neue", size=12)
    # width: all 15, L=50
    for j in range(1, ws.max_column+1):
        col_letter = get_column_letter(j)
        ws.column_dimensions[col_letter].width = 50 if col_letter == "L" else 15
    ws.freeze_panes = "A2"

logging.info(f"✅ Clinician report → {clin_xlsx.name}")

# ─── manifest ────────────────────────────────────────────────────────────
h = hashlib.md5(pd.util.hash_pandas_object(df_clin, index=True).values).hexdigest()
with open(MANIFEST, "a") as mf:
    mf.write(
        f"v8_script3_formatted,{BACKUP_PAR.name},{gold_xlsx.name},{clin_xlsx.name},"
        f"{len(df_clin)} rows,{h},{datetime.datetime.now()}\n"
    )
logging.info("✅ Manifest updated.")